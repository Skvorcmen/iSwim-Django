from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, TemplateView
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.db.models import Q
from django.db import transaction
import openpyxl, io, json
import csv
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from openpyxl.worksheet.datavalidation import DataValidation
from .models import Competition, AgeCategory, Discipline, Registration, Heat, HeatAssignment
from .services import (
    normalize_result_time,
    assign_places_by_time,
    build_heats_for_competition,
    process_application_worksheet,
)
from .query_optimization import get_grouped_heats_optimized
from users.models import AthleteProfile, User as UserModel
from users.auth import secretary_required, has_role
from trainers.models import Trainer

from branches.models import Branch


ALLOWED_STATUS_TRANSITIONS = {
    'upcoming': {'registration', 'closed', 'ongoing', 'cancelled'},
    'registration': {'closed', 'ongoing', 'cancelled'},
    'closed': {'ongoing', 'cancelled'},
    'ongoing': {'finished', 'cancelled'},
    'finished': set(),
    'cancelled': set(),
}


def can_transition_status(current_status, target_status):
    return target_status in ALLOWED_STATUS_TRANSITIONS.get(current_status, set())

def heat_data(request, heat_id):
    heat = get_object_or_404(Heat, id=heat_id)
    assignments = heat.assignments.select_related('registration__athlete__user').order_by('lane')
    data = [{'lane': a.lane, 'name': a.registration.athlete.user.get_full_name(),
             'result_time': a.result_time, 'place': a.place, 
             'id': a.id, 'has_result': bool(a.result_time)} for a in assignments]
    return JsonResponse(data, safe=False)

def competition_api_list(request):
    competitions = Competition.objects.select_related('branch', 'created_by').order_by('start_date')
    data = [
        {
            'id': comp.pk,
            'title': comp.title,
            'location': comp.location,
            'status': comp.status,
            'start_date': comp.start_date.isoformat() if comp.start_date else None,
            'end_date': comp.end_date.isoformat() if comp.end_date else None,
            'branch': comp.branch.name if comp.branch else None,
            'registration_deadline': comp.registration_deadline.isoformat() if comp.registration_deadline else None,
        }
        for comp in competitions
    ]
    return JsonResponse(data, safe=False)

def competition_api_detail(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    data = {
        'id': comp.pk,
        'title': comp.title,
        'description': comp.description,
        'location': comp.location,
        'status': comp.status,
        'start_date': comp.start_date.isoformat() if comp.start_date else None,
        'end_date': comp.end_date.isoformat() if comp.end_date else None,
        'branch': comp.branch.name if comp.branch else None,
        'disciplines': [
            {'id': d.pk, 'style': d.style, 'distance': d.distance}
            for d in comp.disciplines.all()
        ],
        'age_categories': [
            {
                'id': a.pk,
                'name': a.name,
                'birth_year_from': a.birth_year_from,
                'birth_year_to': a.birth_year_to,
                'gender': a.gender,
            }
            for a in comp.age_categories.all()
        ],
    }
    return JsonResponse(data)

def competition_api_results(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    assignments = HeatAssignment.objects.filter(heat__competition=comp).select_related(
        'heat__discipline', 'heat__age_category', 'registration__athlete__user'
    ).order_by('heat__age_category__birth_year_from', 'heat__age_category__gender', 'heat__discipline__style', 'heat__discipline__distance', 'place', 'result_time')
    data = [
        {
            'heat_id': a.heat.pk,
            'discipline': str(a.heat.discipline),
            'category': a.heat.age_category.name,
            'name': a.registration.athlete.user.get_full_name(),
            'result_time': a.result_time,
            'place': a.place,
            'status': a.status,
        }
        for a in assignments
    ]
    return JsonResponse(data, safe=False)

def competition_api_heat_assignments(request, heat_id):
    heat = get_object_or_404(Heat, id=heat_id)
    assignments = heat.assignments.select_related('registration__athlete__user').order_by('lane')
    data = [
        {
            'id': a.id,
            'lane': a.lane,
            'athlete': a.registration.athlete.user.get_full_name(),
            'result_time': a.result_time,
            'place': a.place,
            'status': a.status,
        }
        for a in assignments
    ]
    return JsonResponse(data, safe=False)

class CompetitionListView(ListView):
    model = Competition
    template_name = 'competitions/competition_list.html'
    context_object_name = 'competitions'
    queryset = Competition.objects.all().order_by('start_date')

class CompetitionDetailView(DetailView):
    model = Competition
    template_name = 'competitions/competition_detail.html'
    context_object_name = 'comp'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        from .models import check_registration_deadline
        check_registration_deadline(self.object)
        if request.GET.get('close_registration') and has_role(request.user, 'secretary', 'admin'):
            if can_transition_status(self.object.status, 'closed'):
                self.object.status = 'closed'
                self.object.save(update_fields=['status'])
                messages.success(request, 'Приём заявок закрыт')
            else:
                messages.warning(request, f'Нельзя закрыть приём заявок из статуса «{self.object.get_status_display()}».')
            return redirect('competition_detail', pk=self.object.pk)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['disciplines'] = self.object.disciplines.all()
        ctx['age_categories'] = self.object.age_categories.all()
        ctx['can_close_registration'] = can_transition_status(self.object.status, 'closed')
        if ctx['can_close_registration']:
            ctx['close_registration_reason'] = ''
        else:
            ctx['close_registration_reason'] = f"Недоступно из статуса «{self.object.get_status_display()}»"

        ctx['can_start_live'] = can_transition_status(self.object.status, 'ongoing') or self.object.status == 'ongoing'
        if ctx['can_start_live']:
            ctx['start_live_reason'] = ''
        else:
            ctx['start_live_reason'] = f"Недоступно из статуса «{self.object.get_status_display()}»"

        ctx['trainer_can_apply'] = self.object.status in {'upcoming', 'registration'}
        if ctx['trainer_can_apply']:
            ctx['trainer_apply_reason'] = ''
        else:
            ctx['trainer_apply_reason'] = f"Подача заявок недоступна в статусе «{self.object.get_status_display()}»"

        heats_exist = self.object.heats.exists()
        has_registrations = self.object.registrations.exists()
        all_disciplines_done = False
        if heats_exist:
            groups = get_grouped_heats_optimized(self.object)
            all_disciplines_done = bool(groups) and all(g['is_finished'] for g in groups)

        ctx['status_steps'] = [
            {'label': '1. Приём заявок', 'done': self.object.status in {'closed', 'ongoing', 'finished'}, 'hint': 'Открыт для тренеров'},
            {'label': '2. Заявки собраны', 'done': has_registrations, 'hint': 'Есть загруженные участники'},
            {'label': '3. Заплывы сформированы', 'done': heats_exist, 'hint': 'Созданы heat и дорожки'},
            {'label': '4. Дисциплины завершены', 'done': all_disciplines_done, 'hint': 'Проставлены места по категориям'},
            {'label': '5. Соревнование завершено', 'done': self.object.status == 'finished', 'hint': 'Результаты опубликованы'},
        ]
        done_steps = sum(1 for step in ctx['status_steps'] if step['done'])
        total_steps = len(ctx['status_steps'])
        ctx['status_progress_percent'] = int((done_steps / total_steps) * 100) if total_steps else 0
        ctx['status_progress_text'] = f"{done_steps}/{total_steps}"
        return ctx

class CompetitionCreateView(LoginRequiredMixin, CreateView):
    model = Competition
    template_name = 'competitions/competition_form.html'
    fields = ['title', 'description', 'location', 'branch', 'start_date', 'end_date',
              'registration_deadline', 'lanes', 'regulation', 'image']
    def get_success_url(self):
        return reverse_lazy('competition_detail', kwargs={'pk': self.object.pk})
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.registration_deadline_time = self.request.POST.get("registration_deadline_time", "23:59")
        response = super().form_valid(form)
        cat_names = self.request.POST.getlist('cat_name')
        cat_from = self.request.POST.getlist('cat_from')
        cat_to = self.request.POST.getlist('cat_to')
        cat_gender = self.request.POST.getlist('cat_gender')
        for i in range(len(cat_names)):
            if cat_names[i].strip():
                AgeCategory.objects.create(
                    competition=self.object, name=cat_names[i],
                    birth_year_from=int(cat_from[i]) if cat_from[i] else 2000,
                    birth_year_to=int(cat_to[i]) if cat_to[i] else 2010,
                    gender=cat_gender[i] if i < len(cat_gender) else 'M')
        styles = self.request.POST.getlist('style')
        distances = self.request.POST.getlist('distance')
        for i in range(len(styles)):
            if distances[i]:
                Discipline.objects.get_or_create(
                    competition=self.object, style=styles[i], distance=int(distances[i]))
                # Создаём запись в ленте активности
        from activity.models import Activity
        Activity.objects.create(
            user=self.request.user,
            activity_type='competition',
            title=f'Создано новое соревнование «{self.object.title}»',
            description=f'{self.object.location} | {self.object.start_date}',
            link=f'/competitions/{self.object.pk}/'
        )
        return response

@login_required
def download_template_excel(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявка"
    headers = ['Фамилия', 'Имя', 'Пол (М/Ж)', 'Год рождения', 'Стиль', 'Дистанция (м)', 'Предв. время']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    example = ['Иванов', 'Иван', 'М', 2010, 'Вольный стиль', 50, '00:28,50']
    for col, v in enumerate(example, 1):
        ws.cell(row=2, column=col, value=v)
    dv1 = DataValidation(type="list", formula1='"Вольный стиль,На спине,Брасс,Баттерфляй,Комплекс"')
    ws.add_data_validation(dv1)
    dv1.add('E2:E500')
    dv2 = DataValidation(type="list", formula1='"М,Ж"')
    ws.add_data_validation(dv2)
    dv2.add('C2:C500')
    for col, w in zip(['A','B','C','D','E','F','G'], [15,15,10,12,18,14,18]):
        ws.column_dimensions[col].width = w
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=zayavka_{comp.title.replace(" ", "_")}.xlsx'
    return response

@login_required
def upload_application(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    context = {'comp': comp}
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        coach = request.user
        trainer_profile = getattr(request.user, 'trainer_profile', None)
        branch = trainer_profile.branches.first() if trainer_profile else None
        stats = process_application_worksheet(
            ws=ws,
            comp=comp,
            coach=coach,
            branch=branch,
            user_model=UserModel,
            athlete_profile_model=AthleteProfile,
            discipline_model=Discipline,
            registration_model=Registration,
        )
        messages.success(
            request,
            f"Загружено: {stats['created']}. Дубликатов: {stats['duplicates']}. Пропущено пустых строк: {stats['skipped']}.",
        )
        if stats["errors"]:
            preview = "; ".join(stats["errors"][:3])
            messages.warning(request, f"Ошибки в файле: {preview}")
        request.session[f"import_report_errors_{pk}"] = stats["errors"]
        context['import_report'] = stats
        return render(request, 'competitions/upload_application.html', context)
    return render(request, 'competitions/upload_application.html', context)


@login_required
def download_import_errors_csv(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    errors = request.session.get(f"import_report_errors_{pk}", [])
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="import_errors_{comp.pk}.csv"'
    response.write("\ufeff")
    writer = csv.writer(response, delimiter=";")
    writer.writerow(["competition_id", "competition_title", "error"])
    for error in errors:
        writer.writerow([comp.pk, comp.title, error])
    return response

@secretary_required
def manual_register(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        gender = request.POST.get('gender', 'M')
        birth_year = request.POST.get('birth_year', '2010')
        style = request.POST.get('style', 'free')
        distance = request.POST.get('distance', '50')
        time = request.POST.get('time', '')
        coach_id = request.POST.get('coach_id')
        branch_id = request.POST.get('branch_id')
        if not first_name or not last_name:
            messages.error(request, 'Введите имя и фамилию')
            return redirect('manual_register', pk=pk)
        user, _ = UserModel.objects.get_or_create(
            first_name=first_name, last_name=last_name,
            defaults={'username': f'{last_name}_{first_name}_{birth_year}'.lower()})
        athlete, _ = AthleteProfile.objects.get_or_create(
            user=user, defaults={'birth_date': f'{int(birth_year)}-01-01', 'gender': gender})
        athlete.gender = gender
        athlete.save()
        discipline = Discipline.objects.filter(competition=comp, style=style, distance=int(distance)).first()
        if not discipline:
            discipline, _ = Discipline.objects.get_or_create(competition=comp, style=style, distance=int(distance))
        coach = UserModel.objects.get(id=coach_id) if coach_id else request.user
        branch = Branch.objects.get(id=branch_id) if branch_id else None
        reg, created = Registration.objects.get_or_create(
            competition=comp, athlete=athlete, discipline=discipline,
            defaults={'preliminary_time': time, 'coach': coach, 'branch': branch})
        if created:
            messages.success(request, f'{first_name} {last_name} добавлен в заявки')
        else:
            messages.warning(request, 'Спортсмен уже зарегистрирован на эту дисциплину')
        return redirect('public_results', pk=pk)
    coaches = Trainer.objects.all()
    branches = Branch.objects.filter(is_active=True)
    return render(request, 'competitions/manual_register.html', {'comp': comp, 'coaches': coaches, 'branches': branches})

@secretary_required
def generate_heats(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    comp.heats.filter(assignments__result_time='').delete()
    build_heats_for_competition(
        comp=comp,
        disciplines=comp.disciplines.all(),
        age_categories=comp.age_categories.all(),
        registrations_model=Registration,
        heat_model=Heat,
        assignment_model=HeatAssignment,
    )
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f"competition_{comp.pk}",
        {
            "type": "competition_heats_regenerated",
            "competition_id": comp.pk,
        },
    )
    messages.success(request, 'Заплывы сформированы!')
    return redirect('public_results', pk=pk)


@secretary_required
def manage_heats(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    groups = get_grouped_heats_optimized(comp)
    return render(request, 'competitions/manage_heats.html', {'comp': comp, 'groups': groups})

@secretary_required
def available_athletes(request, heat_id):
    heat = get_object_or_404(Heat, id=heat_id)
    assigned_ids = heat.assignments.values_list('registration__athlete_id', flat=True)
    registrations = Registration.objects.filter(
        competition=heat.competition, discipline=heat.discipline,
        athlete__birth_date__year__gte=heat.age_category.birth_year_from,
        athlete__birth_date__year__lte=heat.age_category.birth_year_to,
        athlete__gender=heat.age_category.gender,
    ).exclude(athlete_id__in=assigned_ids).select_related('athlete__user')
    data = [{'id': r.athlete.id, 'name': r.athlete.user.get_full_name()} for r in registrations]
    return JsonResponse(data, safe=False)

@secretary_required
def add_to_heat(request, heat_id):
    if request.method == 'POST':
        heat = get_object_or_404(Heat, id=heat_id)
        athlete_id = request.POST.get('athlete_id')
        lane = request.POST.get('lane')
        athlete = get_object_or_404(AthleteProfile, id=athlete_id)
        registration = Registration.objects.filter(
            competition=heat.competition, athlete=athlete, discipline=heat.discipline).first()
        if registration:
            HeatAssignment.objects.create(heat=heat, registration=registration, lane=int(lane))
            return JsonResponse({'success': True})
    return JsonResponse({'success': False})

class LiveCompetitionView(LoginRequiredMixin, TemplateView):
    template_name = 'competitions/live_competition.html'
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['comp'] = get_object_or_404(Competition, pk=self.kwargs['pk'])
        if has_role(self.request.user, 'secretary', 'admin') and can_transition_status(ctx['comp'].status, 'ongoing'):
            ctx['comp'].status = 'ongoing'
            ctx['comp'].save(update_fields=['status'])
            messages.success(self.request, 'Соревнование переведено в статус «Идёт сейчас».')
        groups = get_grouped_heats_optimized(ctx['comp'])
        for group in groups:
            if group["is_finished"]:
                group["can_finalize"] = False
                group["finalize_block_reason"] = "Дисциплина уже завершена"
                continue
            has_valid_results = HeatAssignment.objects.filter(
                heat__competition=ctx['comp'],
                heat__discipline=group['discipline'],
                heat__age_category=group['category'],
            ).exclude(result_time='').exists()
            group["can_finalize"] = has_valid_results
            group["finalize_block_reason"] = "" if has_valid_results else "Нет введённых результатов"

        ctx['groups'] = groups
        ctx["all_disciplines_done"] = bool(groups) and all(g["is_finished"] for g in groups)
        if ctx['comp'].status != 'ongoing':
            ctx["can_finish_competition"] = False
            ctx["finish_block_reason"] = f"Статус: {ctx['comp'].get_status_display()}"
        elif not groups:
            ctx["can_finish_competition"] = False
            ctx["finish_block_reason"] = "Нет сформированных заплывов"
        elif not ctx["all_disciplines_done"]:
            ctx["can_finish_competition"] = False
            ctx["finish_block_reason"] = "Не все дисциплины завершены"
        else:
            ctx["can_finish_competition"] = True
            ctx["finish_block_reason"] = ""
        return ctx

@secretary_required
def save_result(request, assignment_id):
    if request.method == 'POST':
        from .websocket_utils import send_delta_result_update
        
        data = json.loads(request.body)
        assignment = get_object_or_404(HeatAssignment, id=assignment_id)
        normalized = normalize_result_time(data.get('time', ''))
        if not normalized:
            return JsonResponse({'success': False, 'error': 'Неверный формат времени. Используйте ММ:СС.сс(с)'}, status=400)
        assignment.result_time = normalized
        assignment.save()
        
        # Отправляем delta-обновление вместо full state
        send_delta_result_update(assignment)
        
        # Места не присваиваем — ждём финализации
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

class PublicResultsView(DetailView):
    model = Competition
    template_name = 'competitions/public_results.html'
    context_object_name = 'comp'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        heats = self.object.heats.all().order_by('age_category__birth_year_from', 'age_category__gender', 'discipline__style', 'discipline__distance', 'number')
        ctx['heats'] = heats
        return ctx


class CompetitionProtocolView(DetailView):
    model = Competition
    template_name = 'competitions/competition_protocol.html'
    context_object_name = 'comp'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        assignments = (
            HeatAssignment.objects.filter(heat__competition=self.object, place__isnull=False)
            .select_related(
                'heat__age_category',
                'heat__discipline',
                'registration__athlete__user',
                'registration__branch',
                'registration__coach',
            )
            .order_by(
                'heat__age_category__birth_year_from',
                'heat__age_category__gender',
                'heat__discipline__style',
                'heat__discipline__distance',
                'place',
                'result_time',
            )
        )
        groups = []
        current_key = None
        for a in assignments:
            key = (a.heat.age_category_id, a.heat.discipline_id)
            if key != current_key:
                current_key = key
                groups.append(
                    {
                        'category': a.heat.age_category,
                        'discipline': a.heat.discipline,
                        'rows': [],
                    }
                )
            groups[-1]['rows'].append(a)
        ctx['protocol_groups'] = groups
        return ctx


def download_competition_protocol_csv(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    assignments = (
        HeatAssignment.objects.filter(heat__competition=comp, place__isnull=False)
        .select_related(
            'heat__age_category',
            'heat__discipline',
            'registration__athlete__user',
            'registration__branch',
            'registration__coach',
        )
        .order_by(
            'heat__age_category__birth_year_from',
            'heat__age_category__gender',
            'heat__discipline__style',
            'heat__discipline__distance',
            'place',
            'result_time',
        )
    )
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="competition_protocol_{comp.pk}.csv"'
    response.write("\ufeff")
    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "competition_id",
        "competition_title",
        "category",
        "gender",
        "discipline",
        "distance",
        "place",
        "athlete",
        "result_time",
        "branch",
        "coach",
    ])
    for row in assignments:
        writer.writerow([
            comp.pk,
            comp.title,
            row.heat.age_category.name,
            row.heat.age_category.get_gender_display(),
            row.heat.discipline.get_style_display(),
            row.heat.discipline.distance,
            row.place,
            row.registration.athlete.user.get_full_name(),
            row.result_time,
            row.registration.branch.name if row.registration.branch else "",
            row.registration.coach.get_full_name() if row.registration.coach else "",
        ])
    return response

@secretary_required
def finalize_results(request, pk, discipline_id, category_id):
    comp = get_object_or_404(Competition, pk=pk)
    if comp.status == 'finished':
        messages.warning(request, 'Соревнование уже завершено. Финализация недоступна.')
        return redirect('live_competition', pk=pk)

    if comp.status not in {'ongoing', 'closed', 'registration', 'upcoming'}:
        messages.warning(request, f'Финализация недоступна в статусе «{comp.get_status_display()}».')
        return redirect('live_competition', pk=pk)

    already_finalized = HeatAssignment.objects.filter(
        heat__competition=comp,
        heat__discipline_id=discipline_id,
        heat__age_category_id=category_id,
        place__isnull=False,
    ).exists()
    if already_finalized:
        messages.info(request, 'Дисциплина уже финализирована.')
        return redirect('live_competition', pk=pk)

    all_assignments = HeatAssignment.objects.filter(
        heat__competition=comp,
        heat__discipline_id=discipline_id,
        heat__age_category_id=category_id,
    ).exclude(result_time='')

    valid_assignments = [a for a in all_assignments if normalize_result_time(a.result_time)]
    if not valid_assignments:
        messages.warning(request, 'Нельзя завершить дисциплину: нет корректно введённых результатов.')
        return redirect('live_competition', pk=pk)

    try:
        from .record_services import check_and_update_records
        
        with transaction.atomic():
            ranked = assign_places_by_time(valid_assignments)
            for assignment in ranked:
                assignment.status = 'finished'
                assignment.save(update_fields=['place', 'status'])
                # Проверяем рекорды
                check_and_update_records(assignment)
            
            # Отмечаем незавершивших как DNF
            dnf = HeatAssignment.objects.filter(
                heat__competition=comp,
                heat__discipline_id=discipline_id,
                heat__age_category_id=category_id,
                result_time=''
            )
            dnf.update(place=None, status='dnf')

    except Exception as e:
        messages.error(request, f'Ошибка при финализации: {str(e)}')
        return redirect('live_competition', pk=pk)

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f"competition_{comp.pk}",
        {
            "type": "competition_discipline_finalized",
            "competition_id": comp.pk,
            "discipline_id": discipline_id,
            "category_id": category_id,
        },
    )
    
    messages.success(request, '✅ Дисциплина завершена! Места распределены. Результаты доступны на странице просмотра.')
    return redirect('public_results', pk=pk)

@secretary_required
def reorder_heats(request, pk):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        for item in data:
            Heat.objects.filter(id=item['id']).update(number=item['number'])
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

@secretary_required
def remove_from_heat(request, assignment_id):
    assignment = get_object_or_404(HeatAssignment, id=assignment_id)
    # Нельзя удалять, если есть результат
    if assignment.result_time:
        return JsonResponse({'success': False, 'error': 'Нельзя удалить участника с результатом'})
    assignment.delete()
    return JsonResponse({'success': True})

def group_heats_data(request, pk):
    cat_id = request.GET.get('cat')
    disc_id = request.GET.get('disc')
    heats = Heat.objects.filter(competition_id=pk, age_category_id=cat_id, discipline_id=disc_id)
    data = []
    for h in heats:
        taken = h.assignments.count()
        free = h.competition.lanes - taken
        data.append({'id': h.id, 'number': h.number, 'taken': taken, 'free': free})
    return JsonResponse(data, safe=False)

def registered_athletes(request, pk):
    q = request.GET.get('q', '').lower()
    regs = Registration.objects.filter(competition_id=pk)
    if q:
        regs = regs.filter(Q(athlete__user__first_name__icontains=q) | Q(athlete__user__last_name__icontains=q))
    data = [{'id': r.athlete.id, 'name': r.athlete.user.get_full_name(), 'discipline': str(r.discipline)} for r in regs[:100]]
    return JsonResponse(data, safe=False)

@secretary_required
def finish_competition(request, pk):
    comp = get_object_or_404(Competition, pk=pk)
    if comp.status == 'finished':
        messages.info(request, 'Соревнование уже завершено.')
        return redirect('public_results', pk=pk)
    if comp.status != 'ongoing':
        messages.warning(request, f'Завершение возможно только в статусе «Идёт сейчас». Текущий статус: «{comp.get_status_display()}».')
        return redirect('live_competition', pk=pk)

    groups = get_grouped_heats_optimized(comp)
    if not groups:
        messages.warning(request, 'Нельзя завершить соревнование без сформированных заплывов.')
        return redirect('live_competition', pk=pk)
    if not all(g['is_finished'] for g in groups):
        messages.warning(request, 'Нельзя завершить соревнование: не все дисциплины финализированы.')
        return redirect('live_competition', pk=pk)

    if not can_transition_status(comp.status, 'finished'):
        messages.warning(request, f'Переход из статуса «{comp.get_status_display()}» в «Завершено» недоступен.')
        return redirect('live_competition', pk=pk)

    comp.status = 'finished'
    comp.save(update_fields=['status'])

    from activity.models import Activity
    from users.models import Achievement
    from .record_services import check_and_update_all_records

    with transaction.atomic():
        # Проверяем и обновляем все рекорды
        new_records = check_and_update_all_records(comp)
        
        # Собираем победителей по каждой дисциплине
        for heat in comp.heats.all():
            for assignment in heat.assignments.filter(place__lte=3):
                athlete = assignment.registration.athlete
                place = assignment.place
                medal = {1: '🥇', 2: '🥈', 3: '🥉'}.get(place, '')

                # Запись в ленту активности
                Activity.objects.get_or_create(
                    user=athlete.user,
                    activity_type='achievement',
                    title=f'{athlete.user.get_full_name()} занял {place} место',
                    link=f'/athlete/{athlete.user.username}/',
                    defaults={
                        'description': f'{heat.discipline.get_style_display()} {heat.discipline.distance}м — {comp.title}',
                    },
                )

                # Создаём достижение
                Achievement.objects.get_or_create(
                    athlete=athlete,
                    title=f'{medal} {place} место — {heat.discipline.get_style_display()} {heat.discipline.distance}м',
                    achievement_type='medal',
                    competition=comp.title,
                    date=comp.start_date,
                    defaults={
                        'description': f'{comp.title} ({comp.start_date})',
                    },
                )

                # Проверка рекорда школы
                Achievement.objects.get_or_create(
                    athlete=athlete,
                    title=f'⭐ Рекорд школы — {heat.discipline.get_style_display()} {heat.discipline.distance}м',
                    achievement_type='record',
                    date=comp.start_date,
                    defaults={
                        'description': f'{assignment.result_time} — {comp.title}',
                    },
                )

        # Общая запись в ленту
        Activity.objects.get_or_create(
            user=request.user,
            activity_type='competition',
            title=f'Соревнование «{comp.title}» завершено!',
            link=f'/competitions/{comp.pk}/results/',
            defaults={
                'description': 'Поздравляем всех участников и победителей!',
            },
        )

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f"competition_{comp.pk}",
        {
            "type": "competition_finished",
            "competition_id": comp.pk,
        },
    )

    messages.success(request, 'Соревнование завершено! Результаты опубликованы.')
    return redirect('public_results', pk=pk)
